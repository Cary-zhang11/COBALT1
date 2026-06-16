#!/usr/bin/env python3
"""
md2excel.py - 将 Markdown 格式的冒烟用例转换为 Excel 文件

用法:
    python md2excel.py <冒烟用例md文件路径> <输出xlsx路径>

依赖: openpyxl
"""

import re
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("错误: 需要安装 openpyxl。请运行: pip install openpyxl")
    sys.exit(1)


def parse_smoke_cases(md_text):
    """解析 Markdown 文本，提取冒烟用例列表"""
    cases = []
    lines = md_text.split('\n')
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # 匹配用例标题: #### SMK-XXX-XX | P0 | 标题
        match = re.match(r'#{3,4}\s+(SMK-[A-Z0-9-]+)\s+\|\s+(P\d+)\s+\|\s+(.+)', line)
        if match:
            case = {
                '编号': match.group(1).strip(),
                '优先级': match.group(2).strip(),
                '标题': match.group(3).strip(),
                '所属模块': '',
                '功能点': '',
                '前置条件': '',
                '测试步骤': '',
                '预期结果': '',
                '三方关联': '',
            }

            i += 1
            # 读取用例的详细字段
            current_field = None
            field_buffer = []

            while i < len(lines):
                sub_line = lines[i]
                stripped = sub_line.strip()

                # 遇到下一个用例标题或新的大章节时停止
                if re.match(r'#{3,4}\s+SMK-', stripped):
                    i -= 1
                    break
                if re.match(r'^#{2}\s+', stripped) and '用例' not in stripped:
                    i -= 1
                    break

                # 匹配字段名
                field_match = re.match(r'- \*\*(所属模块|功能点|前置条件|测试步骤|预期结果|三方关联)\*\*:\s*(.*)', stripped)
                if field_match:
                    # 保存上一个字段的内容
                    if current_field and field_buffer:
                        case[current_field] = '\n'.join(field_buffer).strip()

                    current_field = field_match.group(1)
                    remaining = field_match.group(2).strip()
                    field_buffer = [remaining] if remaining else []
                elif current_field and stripped.startswith('- ') and not stripped.startswith('- **'):
                    # 列表项，属于当前字段
                    field_buffer.append(stripped[2:])
                elif current_field and not stripped:
                    # 空行，可能是字段结束
                    pass
                elif current_field and stripped and not stripped.startswith('#'):
                    # 继续追加到当前字段
                    field_buffer.append(stripped)

                i += 1

            # 保存最后一个字段
            if current_field and field_buffer:
                case[current_field] = '\n'.join(field_buffer).strip()

            cases.append(case)

        i += 1

    return cases


def create_excel(cases, output_path):
    """创建 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "冒烟用例"

    # 定义表头
    headers = ['用例编号', '所属模块', '功能点', '用例标题', '优先级',
               '前置条件', '测试步骤', '预期结果', '三方关联']

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 写入数据
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row_idx, case in enumerate(cases, 2):
        values = [
            case.get('编号', ''),
            case.get('所属模块', ''),
            case.get('功能点', ''),
            case.get('标题', ''),
            case.get('优先级', ''),
            case.get('前置条件', ''),
            case.get('测试步骤', ''),
            case.get('预期结果', ''),
            case.get('三方关联', ''),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # 编号列加粗
            if col_idx == 1:
                cell.font = Font(bold=True)

            # P0 优先级标红
            if col_idx == 5 and value == 'P0':
                cell.font = Font(color='FF0000', bold=True)

    # 设置列宽
    column_widths = [15, 15, 20, 30, 10, 25, 35, 35, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # 设置行高
    for row in range(2, len(cases) + 2):
        ws.row_dimensions[row].height = 60

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    wb.save(output_path)
    print(f"已生成 Excel 文件: {output_path}")
    print(f"共 {len(cases)} 条冒烟用例")


def main():
    if len(sys.argv) < 3:
        print("用法: python md2excel.py <冒烟用例md文件路径> <输出xlsx路径>")
        sys.exit(1)

    md_path = sys.argv[1]
    output_path = sys.argv[2]

    if not os.path.exists(md_path):
        print(f"错误: 文件不存在: {md_path}")
        sys.exit(1)

    with open(md_path, 'r', encoding='utf-8') as f:
        md_text = f.read()

    cases = parse_smoke_cases(md_text)

    if not cases:
        print("警告: 未解析到任何冒烟用例，请检查 Markdown 格式")
        sys.exit(1)

    create_excel(cases, output_path)


if __name__ == '__main__':
    main()
